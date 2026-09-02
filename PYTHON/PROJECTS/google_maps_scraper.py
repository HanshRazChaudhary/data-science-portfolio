"""
Google Maps Local Business Scraper
====================================
Scrapes local business data from Google Maps and saves to Excel.

SETUP:
    pip install selenium openpyxl undetected-chromedriver

USAGE:
    1. Set your search query and location in the CONFIG section below
    2. Run: python google_maps_scraper.py
"""

import time
import re
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException, TimeoutException
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────
#  CONFIG — Edit these before running
# ─────────────────────────────────────────────
SEARCH_QUERY = "dental in lalitpur"        # e.g. "hotels in Thamel", "restaurants in Kathmandu"
MAX_RESULTS  = 100                     # How many businesses to collect
OUTPUT_FILE  = "dental_in_lalitpur.xlsx"
SCROLL_PAUSE = 2                       # Seconds to wait between scrolls
# ─────────────────────────────────────────────


def init_driver():
    options = uc.ChromeOptions()
    options.add_argument("--start-maximized")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    driver = uc.Chrome(options=options)
    return driver


def search_google_maps(driver, query):
    driver.get("https://www.google.com/maps")
    time.sleep(4)

    # Dismiss cookie/consent popup if it appears
    try:
        consent_btn = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.XPATH,
                "//button[contains(., 'Accept') or contains(., 'Agree') or contains(., 'I agree')]"
            ))
        )
        consent_btn.click()
        time.sleep(2)
    except Exception:
        pass

    # Try multiple possible search box selectors
    search_box = None
    for selector in [
        (By.ID, "searchboxinput"),
        (By.XPATH, "//input[@placeholder]"),
        (By.CSS_SELECTOR, "input#searchboxinput"),
        (By.CSS_SELECTOR, "input[name='q']"),
    ]:
        try:
            search_box = WebDriverWait(driver, 8).until(
                EC.presence_of_element_located(selector)
            )
            break
        except Exception:
            continue

    if search_box is None:
        driver.save_screenshot("debug_screenshot.png")
        raise Exception("Search box not found — check debug_screenshot.png")

    search_box.clear()
    search_box.send_keys(query)
    search_box.send_keys(Keys.RETURN)
    time.sleep(4)


def scroll_results(driver, max_results, pause=SCROLL_PAUSE):
    """Scroll the results panel to load more listings."""
    try:
        panel = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "div[role='feed']"))
        )
    except Exception:
        print("⚠️  Could not find results panel.")
        return

    last_count = 0
    for _ in range(50):
        try:
            listings = driver.find_elements(By.CSS_SELECTOR, "div[role='feed'] > div > div[jsaction]")
            count = len(listings)
            if count >= max_results:
                break
            if count == last_count:
                break
            last_count = count
            driver.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight", panel)
            time.sleep(pause)
            print(f"  Loaded {count} listings so far...")
        except StaleElementReferenceException:
            time.sleep(1)
            continue


def safe_find(driver, css_selector, attribute=None):
    """Safely find an element and return its text or attribute, retrying on stale."""
    for _ in range(3):
        try:
            el = driver.find_element(By.CSS_SELECTOR, css_selector)
            if attribute:
                return el.get_attribute(attribute) or ""
            return el.text.strip()
        except StaleElementReferenceException:
            time.sleep(0.5)
        except Exception:
            return ""
    return ""


def extract_listing_data(driver):
    """Extract all data from the currently open listing panel."""
    data = {
        "Name": "",
        "Category": "",
        "Rating": "",
        "Reviews": "",
        "Address": "",
        "Phone": "",
        "Website": "",
        "Hours": "",
        "Google Maps URL": "",
        "Price Range": "",
    }

    # Wait for listing panel to load
    try:
        WebDriverWait(driver, 8).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "h1.DUwDvf, h1[class*='fontHeadlineLarge']"))
        )
    except TimeoutException:
        return data

    # Name
    data["Name"] = safe_find(driver, "h1.DUwDvf, h1[class*='fontHeadlineLarge']")

    # Category
    data["Category"] = safe_find(driver, "button[jsaction*='category'], .DkEaL")

    # Rating
    data["Rating"] = safe_find(driver, "div.F7nice span[aria-hidden='true']")

    # Reviews count
    try:
        for selector in [
            "div.F7nice span[aria-label*='review']",
            "button[jsaction*='review'] span"
        ]:
            try:
                el = driver.find_element(By.CSS_SELECTOR, selector)
                reviews_text = el.get_attribute("aria-label") or el.text
                numbers = re.findall(r"[\d,]+", reviews_text)
                if numbers:
                    data["Reviews"] = numbers[0].replace(",", "")
                    break
            except Exception:
                continue
    except Exception:
        pass

    # Price range
    data["Price Range"] = safe_find(driver, "span[aria-label*='Price']")

    # Address
    data["Address"] = safe_find(driver, "button[data-item-id='address'] div.fontBodyMedium")

    # Phone
    data["Phone"] = safe_find(driver, "button[data-item-id*='phone'] div.fontBodyMedium")

    # Website
    data["Website"] = safe_find(driver, "a[data-item-id='authority']", attribute="href")

    # Hours
    data["Hours"] = safe_find(driver, "div[data-hide-tooltip-on-mouse-out] .fontBodyMedium")

    # Google Maps URL
    try:
        data["Google Maps URL"] = driver.current_url
    except Exception:
        pass

    return data


def get_listing_urls(driver, max_results):
    """Collect all listing URLs first, then visit each one."""
    urls = []
    try:
        panel = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "div[role='feed']"))
        )
    except Exception:
        return urls

    last_count = 0
    for _ in range(50):
        try:
            items = driver.find_elements(By.CSS_SELECTOR, "div[role='feed'] > div > div[jsaction]")
            count = len(items)
            if count >= max_results or count == last_count:
                break
            last_count = count
            driver.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight", panel)
            time.sleep(SCROLL_PAUSE)
            print(f"  Loaded {count} listings so far...")
        except StaleElementReferenceException:
            time.sleep(1)

    # Collect hrefs from all listing anchor tags
    try:
        anchors = driver.find_elements(By.CSS_SELECTOR, "div[role='feed'] a[href*='/maps/place/']")
        seen = set()
        for a in anchors:
            try:
                href = a.get_attribute("href")
                if href and href not in seen:
                    seen.add(href)
                    urls.append(href)
                    if len(urls) >= max_results:
                        break
            except StaleElementReferenceException:
                continue
    except Exception as e:
        print(f"⚠️  Error collecting URLs: {e}")

    return urls


def save_to_excel(businesses, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Local Businesses"

    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", start_color="1A73E8")
    alt_fill     = PatternFill("solid", start_color="EAF2FF")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align   = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border  = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # Title row
    ws.merge_cells("A1:J1")
    ws["A1"] = f"Local Business Data — {SEARCH_QUERY}"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="1A73E8")
    ws["A1"].alignment = center_align
    ws.row_dimensions[1].height = 30

    # Headers
    columns = ["Name", "Category", "Rating", "Reviews", "Address",
               "Phone", "Website", "Hours", "Google Maps URL", "Price Range"]

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = thin_border
    ws.row_dimensions[2].height = 22

    # Data rows
    for row_idx, biz in enumerate(businesses, start=3):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill()
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=biz.get(col_name, ""))
            cell.fill      = fill
            cell.alignment = left_align
            cell.border    = thin_border
            cell.font      = Font(name="Arial", size=10)
        ws.row_dimensions[row_idx].height = 18

    # Column widths
    col_widths = [30, 18, 8, 10, 40, 18, 35, 20, 50, 12]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A3"

    # Summary sheet
    ss = wb.create_sheet("Summary")
    ss["A1"] = "Summary"
    ss["A1"].font = Font(name="Arial", bold=True, size=13)

    summary_data = [
        ("Search Query",    SEARCH_QUERY),
        ("Total Collected", len(businesses)),
        ("With Phone",      sum(1 for b in businesses if b["Phone"])),
        ("With Website",    sum(1 for b in businesses if b["Website"])),
        ("With Rating",     sum(1 for b in businesses if b["Rating"])),
    ]
    for r, (label, value) in enumerate(summary_data, start=3):
        ss.cell(row=r, column=1, value=label).font = Font(name="Arial", bold=True)
        ss.cell(row=r, column=2, value=value)

    ss.column_dimensions["A"].width = 22
    ss.column_dimensions["B"].width = 40

    wb.save(filename)
    print(f"\n✅ Saved {len(businesses)} businesses to '{filename}'")


def main():
    print(f"🔍 Searching: {SEARCH_QUERY}")
    print(f"🎯 Target: {MAX_RESULTS} results\n")

    driver = init_driver()
    businesses = []

    try:
        search_google_maps(driver, SEARCH_QUERY)

        print("📜 Collecting listing URLs...")
        urls = get_listing_urls(driver, MAX_RESULTS)
        print(f"\n📦 Found {len(urls)} listings. Extracting details...\n")

        for i, url in enumerate(urls, start=1):
            try:
                driver.get(url)
                time.sleep(2.5)

                data = extract_listing_data(driver)
                name = data["Name"] or url[:50]
                print(f"  [{i}/{len(urls)}] {name[:50]}")

                if data["Name"]:
                    businesses.append(data)

                # Save progress every 20 businesses
                if len(businesses) % 20 == 0 and businesses:
                    save_to_excel(businesses, OUTPUT_FILE)
                    print(f"  💾 Progress saved ({len(businesses)} so far)...")

            except Exception as e:
                print(f"  ⚠️  Skipped listing {i}: {e}")
                continue

    finally:
        driver.quit()

    if businesses:
        save_to_excel(businesses, OUTPUT_FILE)
        print(f"\n📊 Final Summary:")
        print(f"   Total scraped  : {len(businesses)}")
        print(f"   With phone     : {sum(1 for b in businesses if b['Phone'])}")
        print(f"   With website   : {sum(1 for b in businesses if b['Website'])}")
        print(f"   With rating    : {sum(1 for b in businesses if b['Rating'])}")
    else:
        print("\n❌ No businesses found. Try adjusting SEARCH_QUERY.")


if __name__ == "__main__":
    main()